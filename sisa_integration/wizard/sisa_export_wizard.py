# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
import io
import csv
from datetime import datetime
import logging

_logger = logging.getLogger(__name__)


class SisaExportWizard(models.TransientModel):
    _name = 'sisa.export.wizard'
    _description = 'Asistente de Exportación SISA'

    declaration_id = fields.Many2one(
        'sisa.declaration',
        string='Declaración SISA',
        required=True,
        ondelete='cascade'
    )
    
    export_format = fields.Selection([
        ('csv', 'CSV (Separado por comas)'),
        ('txt', 'TXT (Separado por tabulaciones)'),
        ('xlsx', 'Excel (XLSX)')
    ], string='Formato de Exportación', default='csv', required=True)
    
    include_header = fields.Boolean(
        string='Incluir Encabezados',
        default=True,
        help="Incluir fila de encabezados en el archivo exportado"
    )
    
    separator = fields.Selection([
        (',', 'Coma (,)'),
        (';', 'Punto y coma (;)'),
        ('tab', 'Tabulación'),
        ('|', 'Pipe (|)')
    ], string='Separador', default=',',
       help="Separador de campos para archivos CSV/TXT")
    
    encoding = fields.Selection([
        ('utf-8', 'UTF-8'),
        ('latin-1', 'Latin-1 (ISO 8859-1)'),
        ('cp1252', 'Windows-1252')
    ], string='Codificación', default='utf-8',
       help="Codificación de caracteres del archivo")
    
    # Campos para configurar qué exportar
    export_stock = fields.Boolean(
        string='Exportar Existencias',
        default=True,
        help="Incluir existencias de granos en la exportación"
    )
    
    export_surface = fields.Boolean(
        string='Exportar Superficie',
        default=True,
        help="Incluir superficie sembrada en la exportación"
    )
    
    # Campos de resultado
    file_data = fields.Binary(
        string='Archivo Generado',
        readonly=True
    )
    
    file_name = fields.Char(
        string='Nombre del Archivo',
        readonly=True
    )
    
    state = fields.Selection([
        ('init', 'Configuración'),
        ('done', 'Completado')
    ], default='init', string='Estado')
    
    # Información de la declaración
    declaration_info = fields.Html(
        string='Información de la Declaración',
        compute='_compute_declaration_info'
    )

    @api.depends('declaration_id')
    def _compute_declaration_info(self):
        """Compute declaration information"""
        for wizard in self:
            if wizard.declaration_id:
                decl = wizard.declaration_id
                type_name = dict(decl._fields['declaration_type'].selection)[decl.declaration_type]
                
                info = f"""
                <div class="alert alert-info">
                    <strong>{type_name} - {decl.year}</strong><br/>
                    Estado: <span class="badge badge-{decl.state}">{dict(decl._fields['state'].selection)[decl.state]}</span><br/>
                    Fecha de Generación: {decl.generation_date.strftime('%d/%m/%Y %H:%M')}<br/>
                """
                
                if decl.declaration_type == 'ip1':
                    info += f"""
                    Total Stock: {decl.total_stock_kg:,.0f} kg<br/>
                    Total Superficie: {decl.total_surface_ha:,.1f} ha<br/>
                    Líneas de Stock: {len(decl.stock_line_ids)}<br/>
                    Líneas de Superficie: {len(decl.surface_line_ids)}
                    """
                else:  # IP2
                    info += f"""
                    Total Superficie: {decl.total_surface_ha:,.1f} ha<br/>
                    Líneas de Superficie: {len(decl.surface_line_ids)}
                    """
                
                info += "</div>"
                wizard.declaration_info = info
            else:
                wizard.declaration_info = ""

    @api.onchange('declaration_id')
    def _onchange_declaration_id(self):
        """Configure export options based on declaration type"""
        if self.declaration_id:
            if self.declaration_id.declaration_type == 'ip1':
                self.export_stock = True
                self.export_surface = True
            else:  # IP2
                self.export_stock = False
                self.export_surface = True

    def action_generate_file(self):
        """Generate the export file"""
        self.ensure_one()
        
        if not self.declaration_id:
            raise UserError("Debe seleccionar una declaración")
        
        if not (self.export_stock or self.export_surface):
            raise UserError("Debe seleccionar al menos una opción de exportación")
        
        # Generate file based on format
        if self.export_format in ('csv', 'txt'):
            file_data, file_name = self._generate_csv_file()
        elif self.export_format == 'xlsx':
            file_data, file_name = self._generate_xlsx_file()
        else:
            raise UserError("Formato de exportación no soportado")
        
        # Update declaration export info
        self.declaration_id.write({
            'exported_file_name': file_name,
            'export_date': fields.Datetime.now()
        })
        
        self.file_data = file_data
        self.file_name = file_name
        self.state = 'done'
        
        return self._return_wizard_action()

    def _generate_csv_file(self):
        """Generate CSV/TXT file"""
        output = io.StringIO()
        
        # Get separator
        separator = '\t' if self.separator == 'tab' else self.separator
        writer = csv.writer(output, delimiter=separator, quoting=csv.QUOTE_MINIMAL)
        
        # Determine file extension
        ext = 'txt' if self.export_format == 'txt' else 'csv'
        
        # Generate file name
        decl = self.declaration_id
        type_short = 'IP1' if decl.declaration_type == 'ip1' else 'IP2'
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = f"SISA_{type_short}_{decl.year}_{timestamp}.{ext}"
        
        # Export stock lines (IP1 only)
        if self.export_stock and decl.declaration_type == 'ip1' and decl.stock_line_ids:
            if self.include_header:
                writer.writerow([
                    'TIPO', 'PRODUCTO_CODIGO', 'PRODUCTO_NOMBRE', 
                    'UBICACION_CODIGO', 'UBICACION_NOMBRE', 'CANTIDAD_KG'
                ])
            
            for line in decl.stock_line_ids:
                writer.writerow([
                    'STOCK',
                    line.product_id.default_code or '',
                    line.product_id.name,
                    line.location_id.barcode or line.location_id.id,
                    line.location_id.complete_name,
                    line.quantity_kg
                ])
        
        # Export surface lines
        if self.export_surface and decl.surface_line_ids:
            # Add separator line if we already wrote stock data
            if self.export_stock and decl.declaration_type == 'ip1' and decl.stock_line_ids:
                writer.writerow([])  # Empty line
            
            if self.include_header:
                writer.writerow([
                    'TIPO', 'CULTIVO_CODIGO', 'CULTIVO_NOMBRE',
                    'CAMPO_NOMBRE', 'LOTE_NOMBRE', 'PARTIDA_INMOBILIARIA',
                    'SUPERFICIE_HA', 'FECHA_SIEMBRA'
                ])
            
            for line in decl.surface_line_ids:
                writer.writerow([
                    'SUPERFICIE',
                    line.crop_id.default_code or '',
                    line.crop_id.name,
                    line.field_id.name,
                    line.lot_id.name,
                    line.real_estate_id or '',
                    line.area,
                    line.planting_date.strftime('%d/%m/%Y') if line.planting_date else ''
                ])
        
        # Get content and encode
        content = output.getvalue()
        output.close()
        
        # Encode to bytes
        file_data = base64.b64encode(content.encode(self.encoding))
        
        return file_data, file_name

    def _generate_xlsx_file(self):
        """Generate Excel file"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise UserError(
                "La librería openpyxl no está instalada. "
                "Use formato CSV o TXT como alternativa."
            )
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        decl = self.declaration_id
        
        # Create stock sheet (IP1 only)
        if self.export_stock and decl.declaration_type == 'ip1' and decl.stock_line_ids:
            ws_stock = wb.create_sheet("Existencias")
            
            # Header style
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Headers
            headers = ['Producto Código', 'Producto Nombre', 'Ubicación Código', 
                      'Ubicación Nombre', 'Cantidad (kg)']
            for col, header in enumerate(headers, 1):
                cell = ws_stock.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data
            for row, line in enumerate(decl.stock_line_ids, 2):
                ws_stock.cell(row=row, column=1, value=line.product_id.default_code or '')
                ws_stock.cell(row=row, column=2, value=line.product_id.name)
                ws_stock.cell(row=row, column=3, value=line.location_id.barcode or str(line.location_id.id))
                ws_stock.cell(row=row, column=4, value=line.location_id.complete_name)
                ws_stock.cell(row=row, column=5, value=line.quantity_kg)
        
        # Create surface sheet
        if self.export_surface and decl.surface_line_ids:
            ws_surface = wb.create_sheet("Superficie")
            
            # Header style
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Headers
            headers = ['Cultivo Código', 'Cultivo Nombre', 'Campo', 'Lote', 
                      'Partida Inmobiliaria', 'Superficie (ha)', 'Fecha Siembra']
            for col, header in enumerate(headers, 1):
                cell = ws_surface.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data
            for row, line in enumerate(decl.surface_line_ids, 2):
                ws_surface.cell(row=row, column=1, value=line.crop_id.default_code or '')
                ws_surface.cell(row=row, column=2, value=line.crop_id.name)
                ws_surface.cell(row=row, column=3, value=line.field_id.name)
                ws_surface.cell(row=row, column=4, value=line.lot_id.name)
                ws_surface.cell(row=row, column=5, value=line.real_estate_id or '')
                ws_surface.cell(row=row, column=6, value=line.area)
                ws_surface.cell(row=row, column=7, value=line.planting_date.strftime('%d/%m/%Y') if line.planting_date else '')
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        
        # Generate file name
        type_short = 'IP1' if decl.declaration_type == 'ip1' else 'IP2'
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = f"SISA_{type_short}_{decl.year}_{timestamp}.xlsx"
        
        file_data = base64.b64encode(output.getvalue())
        output.close()
        
        return file_data, file_name

    def action_download_file(self):
        """Download generated file"""
        self.ensure_one()
        if not self.file_data:
            raise UserError("No hay archivo generado")
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content?model=sisa.export.wizard&id={self.id}&field=file_data&download=true&filename={self.file_name}',
            'target': 'self'
        }

    def _return_wizard_action(self):
        """Return action to keep wizard open"""
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sisa.export.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new'
        }